import os
import openpyxl
import json
mod_id = "Forwindz.CustomCornerstones"



workbook = openpyxl.load_workbook('./input/⑨の自定义基石表_mod.xlsx',read_only=True,data_only=True)
sheet = workbook['基石-语言文件信息'] 


sequenceNameType=["displayName","description","preview","retroactive","other_1","other_2"]
langs =["zh-CN","en"]

contents={}

for lang in langs:
    contents[lang] = {}

for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row): 
    programName = row[0].value
    print(programName,"----------------------")
    for i,cell in enumerate(row):
        if i==0:
            continue
        typeIndex = (i-1)//2
        langIndex = i%2
        print(i," > ",typeIndex,langIndex)
        typeName = sequenceNameType[typeIndex]
        langName = langs[langIndex]
        localizationName = "%s_%s_%s"%(mod_id,programName,typeName)
        v = cell.value
        print(i," > ",langName,localizationName, v)
        if v is not None:
            contents[langName][localizationName]=v

for lang in langs:
    filePath = os.path.join("../Assets/lang/","%s.json"%lang)
    with open(filePath,"w",encoding="utf-8") as f:
        json.dump(contents[lang],f,indent=4)